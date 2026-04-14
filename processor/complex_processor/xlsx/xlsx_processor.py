import csv
import io
from dataclasses import dataclass, field
from pathlib import Path

from fast_ai.exceptions import ConfigurationError
from processor.complex_processor.base_processor import BaseComplexProcessor


@dataclass
class Sheet:
    name: str
    csv: str


@dataclass
class XLSX:
    sheets: list[Sheet] = field(default_factory=list)

    @property
    def result(self) -> str:
        return "\n\n".join(f"{s.name}:{s.csv}" for s in self.sheets)


class XLSXProcessor(BaseComplexProcessor):
    """Converts each sheet of an ``.xlsx`` workbook to CSV.

    Config format (``xlsx_processor_config.yaml``)::

        csv_delimiter: ","
    """

    def __init__(self, csv_delimiter: str = ","):
        self.csv_delimiter = csv_delimiter

    @classmethod
    def build(cls, config_path: str | Path | None = None) -> "XLSXProcessor":
        if config_path is None:
            return cls()

        import yaml

        path = Path(config_path)
        if not path.exists():
            raise ConfigurationError(f"config file not found: {path}")

        with open(path, encoding="utf-8") as f:
            cfg = yaml.safe_load(f) or {}

        return cls(csv_delimiter=str(cfg.get("csv_delimiter", ",")))

    def run(self, source: str | Path | bytes) -> XLSX:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ConfigurationError(
                "openpyxl is required for XLSXProcessor. "
                "Install it with `pip install openpyxl`.",
            ) from exc

        if isinstance(source, bytes):
            wb = load_workbook(filename=io.BytesIO(source), read_only=True, data_only=True)
        else:
            path = Path(source)
            if not path.exists():
                raise FileNotFoundError(f"xlsx not found: {path}")
            wb = load_workbook(filename=path, read_only=True, data_only=True)

        try:
            sheets: list[Sheet] = []
            for ws in wb.worksheets:
                sheets.append(Sheet(name=ws.title, csv=self._sheet_to_csv(ws)))
            return XLSX(sheets=sheets)
        finally:
            wb.close()

    def _sheet_to_csv(self, ws) -> str:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=self.csv_delimiter, lineterminator="\n")
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
        return buf.getvalue()
